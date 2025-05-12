
import streamlit as st
import pandas as pd
import math
import io
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ──────────────── Front‐end Configuration ────────────────
st.set_page_config(page_title="Exam Seating Arrangement", layout="wide")
st.title("Exam Seating Arrangement Generator Project IIT Patna Created Using Python By Anuj raj")

uploaded = st.file_uploader("Upload input Excel", type="xlsx")
if not uploaded:
    st.info("Please upload the Excel file.")
    st.stop()

# ──────────────── 2) READ SHEETS ────────────────
xls      = pd.ExcelFile(uploaded)
df_tt    = xls.parse("in_timetable")
df_cr    = xls.parse("in_course_roll_mapping")
df_rn    = xls.parse("in_roll_name_mapping")
df_room  = xls.parse("in_room_capacity")

# ──────────────── 3) INPUTS ────────────────
buffer   = st.number_input("Buffer seats per room", min_value=0, step=1, value=5)
density  = st.radio("Seating Density", ["Sparse", "Dense"])

# ──────────────── 4) Prepare mappings ────────────────
df_cr.columns         = df_cr.columns.str.strip()
df_cr['course_code']  = df_cr['course_code'].str.upper().str.strip()
df_cr['rollno']       = df_cr['rollno'].str.upper().str.strip()
df_rn.columns         = df_rn.columns.str.strip()
df_rn['Roll']         = df_rn['Roll'].str.upper().str.strip()

name_dict      = pd.Series(df_rn.Name.values, index=df_rn.Roll).to_dict()
course_to_rolls = df_cr.groupby('course_code')['rollno'] \
                       .apply(lambda s: sorted(s.tolist())) \
                       .to_dict()

rooms = []
for _, r in df_room.iterrows():
    rid   = str(r['Room No.']).strip()
    cap   = int(r['Exam Capacity'])
    block = str(r['Block']).strip()
    num   = int(rid) if block == 'B1' else int(rid.split('-')[-1])
    rooms.append(dict(room=rid, capacity=cap, block=block, numeric=num))

# ──────────────── 5) Allocation logic ────────────────
def allocate_course(rolls, avail):
    N = len(rolls)
    allowed = []
    for r in avail:
        eff = r['capacity'] - buffer
        if eff <= 0: continue
        use = math.floor(eff*0.5) if density == 'Sparse' else eff
        if use > 0:
            allowed.append(dict(**r, allowed=use))
    if sum(r['allowed'] for r in allowed) < N:
        return None

    # Attempt single‐block B1 then B2
    for blk in ('B1','B2'):
        pool = sorted([r for r in allowed if r['block'] == blk], key=lambda x: x['numeric'])
        if sum(r['allowed'] for r in pool) >= N:
            alloc, rem, idx = [], N, 0
            for rr in pool:
                if rem <= 0: break
                t = min(rem, rr['allowed'])
                alloc.append((rr['room'], rolls[idx:idx+t]))
                rem -= t; idx += t
            return alloc

    # Otherwise split across both blocks
    allowed.sort(key=lambda x: -x['allowed'])
    alloc, rem, idx = [], N, 0
    for rr in allowed:
        if rem <= 0: break
        t = min(rem, rr['allowed'])
        alloc.append((rr['room'], rolls[idx:idx+t]))
        rem -= t; idx += t
    return None if rem > 0 else alloc

# ──────────────── 6) Build data structures ────────────────
overall    = []
seats_left = []
per_date   = {}

for _, row in df_tt.iterrows():
    date = pd.to_datetime(row['Date']).strftime("%d_%m_%Y")
    per_date.setdefault(date, {'morning':{}, 'evening':{}})

    morn = [] if pd.isna(row['Morning']) else [c.strip() for c in row['Morning'].split(';')]
    eve  = [] if pd.isna(row['Evening']) else [c.strip() for c in row['Evening'].split(';')]

    # Morning
    avail, overflow = rooms.copy(), []
    for c in sorted(morn, key=lambda x: -len(course_to_rolls.get(x,[]))):
        rolls = course_to_rolls.get(c, [])
        alloc = allocate_course(rolls, avail)
        if not alloc:
            overflow.append(c)
        else:
            for room, grp in alloc:
                avail = [r for r in avail if r['room'] != room]
                per_date[date]['morning'][(c,room)] = grp
                overall.append({'Date':date,'Course':c,'Room':room,'Rolls':";".join(grp)})

    # Evening
    avail = rooms.copy()
    for c in sorted(eve + overflow, key=lambda x: -len(course_to_rolls.get(x,[]))):
        rolls = course_to_rolls.get(c, [])
        alloc = allocate_course(rolls, avail)
        if not alloc:
            seats_left.append({'Date':date,'Course':c,'Unallocated':len(rolls)})
        else:
            for room, grp in alloc:
                avail = [r for r in avail if r['room'] != room]
                per_date[date]['evening'][(c,room)] = grp
                overall.append({'Date':date,'Course':c,'Room':room,'Rolls':";".join(grp)})

# ──────────────── 7) Show tables ────────────────
st.subheader("Overall Seating")
st.dataframe(pd.DataFrame(overall))

st.subheader("Seats Left")
st.dataframe(pd.DataFrame(seats_left))

# ──────────────── 8) Styling helper ────────────────
thin = Side(style='thin')
def style_ws(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border    = Border(thin,thin,thin,thin)
            cell.alignment = Alignment('center','center')
    for col in ws.columns:
        w = max(len(str(c.value)) if c.value else 0 for c in col) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = w

# ──────────────── 9) Build ZIP ────────────────
zip_buf = io.BytesIO()
with zipfile.ZipFile(zip_buf, "w") as z:
    # overall_seating.xlsx
    def to_xlsx_bytes(df, sheet_name):
        b = io.BytesIO()
        with pd.ExcelWriter(b, engine='openpyxl') as w:
            df.to_excel(w, sheet_name=sheet_name, index=False)
        return b.getvalue()

    z.writestr("overall_seating.xlsx", to_xlsx_bytes(pd.DataFrame(overall),"Overall"))
    z.writestr("seats_left.xlsx",    to_xlsx_bytes(pd.DataFrame(seats_left),"SeatsLeft"))

    # per-date folders
    for date, sessions in per_date.items():
        for sess in ("morning","evening"):
            prefix = f"{date}/{sess}/"
            for (course,room), grp in sessions[sess].items():
                wb = Workbook(); ws = wb.active; ws.title = sess
                ws.append([f"Course: {course} | Room: {room} | Date: {date.replace('_','-')} | Session: {sess.capitalize()}"])
                ws.append(["Roll","Student Name","Signature"])
                for rn in grp:
                    ws.append([rn, name_dict.get(rn,"Unknown Name"), ""])
                # TA & Inv rows
                ws.append([])
                for i in range(1,6): ws.append([f"TA{i}","",""])
                ws.append([])
                for i in range(1,6): ws.append([f"Invigilator{i}","",""])
                style_ws(ws)

                out   = io.BytesIO()
                wb.save(out)
                data  = out.getvalue()
                fname = f"{prefix}{date}_{course}_{room}_{sess}.xlsx"
                z.writestr(fname, data)

zip_buf.seek(0)
st.download_button(
    " Download Full ZIP",
    data=zip_buf,
    file_name="schedules.zip",
    mime="application/zip"
)

st.success(" schedules.zip is ready! Click to download. Regards Anujraj")
